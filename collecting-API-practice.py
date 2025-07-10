import os 
import requests
from PIL import Image
from IPython.display import IFrame
from openpyxl import Workbook

## Write your code here
# URL = 'https://cf-courses-data.s3.us.cloud-object-storage.appdomain.cloud/IBMDeveloperSkillsNetwork-PY0101EN-SkillsNetwork/labs/Module%205/data/Example1.txt'
# r1=requests.get(URL)
# path=os.path.join(os.getcwd(),'text-url.txt')
# with open("text-url.txt","wb") as f:
#     f.write(r1.content)

# url_get='http://httpbin.org/get'
# payload={"name":"Joseph","ID":"123"}
# r2=requests.get(url_get,params=payload)
# print(r2.json())

api_url="https://cf-courses-data.s3.us.cloud-object-storage.appdomain.cloud/IBM-DA0321EN-SkillsNetwork/labs/module%201/Accessing%20Data%20Using%20APIs/jobs.json"

r=requests.get(api_url)
if r.ok:
    print("Information successfully requested")
    data=r.json()
    print("Information loaded into variable.")

def get_number_jobs_t(technology):
    number=0
    for id in data:
        if id.get("Key Skills")==technology:
            number+=1
    return technology, number

def get_number_jobs_l(location):
    number=0
    for id in data:
        if id.get("Location")==location:
            number+=1
    return location, number

def save_results():
    wb=Workbook()
    ws1 = wb.create_sheet("Technology")
    ws1.append(['Technology','Number of Jobs'])
    for tech in tech_list:
        t,n=get_number_jobs_t(tech)
        ws1.append([t,n])
    ws2 = wb.create_sheet("Location")
    ws2.append(['Location','Number of Jobs'])
    for loc in loc_list:
        t,n=get_number_jobs_l(loc)
        ws2.append([t,n])
    wb.save("job-postings.xlsx")
    return 0 

tech_list=[
    "C",
    "C#",
    "C++",
    "Java",
    "JavaScript",
    "Python",
    "Scala",
    "Oracle",
    "SQL Server",
    "MySQL Server",
    "PostgreSQL",
    "MongoDB"
]
loc_list=[
    "Los Angeles",
    "New York",
    "San Francisco",
    "Washington DC",
    "Seattle",
    "Austin",
    "Detroit"
]

save_results()

